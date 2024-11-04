import os
import warnings

from openpyxl.reader.excel import load_workbook as load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment


class Statistics:
    """
    Keeps track of changed and cancelled lessons. Use separate instances for each timetable to track.

    **data sheet columns**: timestamp, planned_teacher, planned_subject, actual_teacher, actual_subject,
    is_cancelled, comment

    **statistics sheet columns**: percentage_changed_teacher, percentage_changed_subject, percentage_cancelled
    """

    def __init__(self, filename, title):
        self.filename = filename
        self.workbook_filename = None
        self.title = title
        self.data = dict()
        self.count_all = 0
        self.count_changed_teacher = 0
        self.count_changed_subject = 0
        self.count_cancelled = 0

    def open(self):
        """
        Reads the existing data, if any.
        """
        warnings.filterwarnings(action="ignore",
                                message="Title is more than 31 characters. Some applications may not be able to read "
                                        "the file",
                                module="openpyxl")

        self.workbook_filename = os.path.realpath(self.filename)
        if os.path.isfile(self.workbook_filename):
            workbook = load_workbook(self.workbook_filename)
            datasheet = workbook[self.title]
            if datasheet is not None:
                for row in datasheet.iter_rows(min_row=2, max_col=7, values_only=True):
                    if row[0] is None:
                        break
                    self.data[row[0]] = {"planned_teacher": row[1],
                                         "planned_subject": row[2],
                                         "actual_teacher": row[3],
                                         "actual_subject": row[4],
                                         "is_cancelled": row[5],
                                         "comment": row[6]}

    def save(self):
        workbook = Workbook()
        workbook.remove_sheet(workbook.active)
        datasheet = workbook.create_sheet(self.title)
        datasheet.append(["timestamp", "planned_teacher", "planned_subject", "actual_teacher", "actual_subject",
                          "is_cancelled", "comment"])
        statsheet = workbook.create_sheet(f"{self.title} - Statistics")
        statsheet.append(["percentage_changed_teacher", "percentage_changed_subject", "percentage_cancelled"])

        self.count_all = 0
        self.count_changed_teacher = 0
        self.count_changed_subject = 0
        self.count_cancelled = 0

        for timestamp in sorted(self.data.keys()):
            entry = self.data[timestamp]
            datasheet.append([timestamp, entry["planned_teacher"], entry["planned_subject"], entry["actual_teacher"],
                              entry["actual_subject"], entry["is_cancelled"], entry["comment"]])
            self.count_all += 1
            if entry["is_cancelled"]:
                self.count_cancelled += 1
            else:
                if entry["planned_teacher"] != entry["actual_teacher"]:
                    self.count_changed_teacher += 1
                if entry["planned_subject"] != entry["actual_subject"]:
                    self.count_changed_subject += 1
        for cell in datasheet["A"]:
            cell.alignment = Alignment(horizontal='left')
            cell.number_format = 'YYYY-MM-DD HH:MM:SS'
        for column_cells in datasheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            datasheet.column_dimensions[column_cells[0].column_letter].width = length

        statsheet.append([self.count_changed_teacher / self.count_all,
                          self.count_changed_subject / self.count_all,
                          self.count_cancelled / self.count_all])
        statsheet["A2"].number_format = '0.00" "%'
        statsheet["B2"].number_format = '0.00" "%'
        statsheet["C2"].number_format = '0.00" "%'
        for column_cells in statsheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells) + 3
            statsheet.column_dimensions[column_cells[0].column_letter].width = length
        workbook.save(os.path.realpath(self.filename))

    def put(self, timestamp, planned_teacher, planned_subject, actual_teacher=None, actual_subject=None,
            is_cancelled=None, comment=None):
        if planned_teacher or planned_subject or actual_teacher or actual_subject:
            self.data[timestamp] = {"planned_teacher": planned_teacher,
                                    "planned_subject": planned_subject,
                                    "actual_teacher": actual_teacher,
                                    "actual_subject": actual_subject,
                                    "is_cancelled": is_cancelled,
                                    "comment": comment}

    def earliest_date(self):
        return sorted(self.data.keys())[0]

    def percentage_cancelled(self):
        return self.count_cancelled / self.count_all

    def percentage_changed_subject(self):
        return self.count_changed_subject / self.count_all

    def percentage_changed_teacher(self):
        return self.count_changed_teacher / self.count_all
